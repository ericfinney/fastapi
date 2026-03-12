import os
import uuid
import json
import logging
import re
import time
from copy import copy
from typing import Dict, Any, List, Optional, Tuple

from fastapi import FastAPI, Body, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Protection, Font

from pydantic import BaseModel
from pypdf import PdfReader
import base64
from io import BytesIO
from pathlib import Path

logging.basicConfig(level=logging.INFO)

TEMPLATE_PATH = os.environ.get("BOYD_TEMPLATE_PATH", "templates/Blank.xlsx")
SHEET_NAME = os.environ.get("BOYD_SHEET_NAME", "Proposal")
OUTPUT_DIR = os.environ.get("BOYD_OUTPUT_DIR", "/tmp/output")
LOGO_PATH = os.environ.get("BOYD_LOGO_PATH", "assets/logo.png")

# Fallback mode for big PDFs:
MAX_PDF_PAGES_STREAMING = int(os.environ.get("BOYD_MAX_PDF_PAGES_STREAMING", "40"))
MAX_EXTRACTED_TEXT_CHARS = int(os.environ.get("BOYD_MAX_EXTRACTED_TEXT_CHARS", "250000"))

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ✅ FIX: remove stray "f"
app = FastAPI()

# Treat anything smaller than half a cent as zero (guards rounding noise)
SUBTOTAL_EPSILON = 0.005

SUBTOTAL_ROW_RE = re.compile(r'^\s*(Shipping|Install|Permitting)\s+Subtotal\s+\$', re.I)
SIGN_PACKAGE_HEADER_RE_EXCEL = re.compile(r'\bPACKAGE\s*$', re.I)  # Match any line ending with PACKAGE

# Strict sign code: letter '.' digit ... (prevents BLDG. from matching)
SIGN_CODE_STRICT_RE = re.compile(r'^\s*([A-Za-z]\.(?:\d+[A-Za-z0-9]*)(?:\.\d+[A-Za-z0-9]*)*)\b')
# Non-anchored strict sign code (for finding embedded codes inside a merged description)
SIGN_CODE_STRICT_ANYWHERE_RE = re.compile(r'([A-Za-z]\.(?:\d+[A-Za-z0-9]*)(?:\.\d+[A-Za-z0-9]*)*)\s*[-–—]')


# =========================================================
# Cleanup: delete old generated workbooks
# =========================================================
def cleanup_old_generated_workbooks(
    directory: str,
    older_than_minutes: int = 30,
    filename_prefix: str = "Boyd_Proposal_",
    allowed_exts: Tuple[str, ...] = (".xlsx", ".xlsm"),
) -> None:
    if not os.path.isdir(directory):
        return
    now = time.time()
    max_age_sec = older_than_minutes * 60

    for fn in os.listdir(directory):
        fn_lower = fn.lower()
        if not fn.startswith(filename_prefix):
            continue
        if not fn_lower.endswith(tuple(ext.lower() for ext in allowed_exts)):
            continue

        path = os.path.join(directory, fn)
        if not os.path.isfile(path):
            continue

        try:
            age_sec = now - os.path.getmtime(path)
            if age_sec > max_age_sec:
                os.remove(path)
                logging.info("Deleted old output workbook: %s (age %.1f min)", fn, age_sec / 60.0)
        except Exception as e:
            logging.warning("Failed deleting %s: %s", fn, e)


# =========================================================
# Basic helpers
# =========================================================
def safe_str(x) -> str:
    return "" if x is None else str(x)

def safe_num(x):
    try:
        if x is None or x == "":
            return None
        return float(x)
    except Exception:
        return None

def round_nearest_dollar(value):
    if value is None or value == "":
        return None
    try:
        return round(float(value))
    except Exception:
        return None

def join_address_lines(addr_lines: List[str]) -> str:
    return "\n".join([line for line in addr_lines if line and line.strip()])

def write_cell(ws, cell: str, value):
    ws[cell].value = value

def insert_logo(ws):
    if not os.path.exists(LOGO_PATH):
        logging.warning("Logo not found at %s; skipping insert.", LOGO_PATH)
        return
    img = XLImage(LOGO_PATH)
    ws.add_image(img, "A1")

def is_bold_description_row(text: str) -> bool:
    t = (text or "").strip()
    if not t:
        return False
    return bool(SIGN_PACKAGE_HEADER_RE_EXCEL.search(t) or SUBTOTAL_ROW_RE.search(t))

def extract_sign_code_strict(raw: str) -> str:
    if not raw:
        return ""
    m = SIGN_CODE_STRICT_RE.search(str(raw))
    return m.group(1).strip() if m else ""

FALLBACK_CODE_RE = re.compile(r'^\s*([A-Za-z]\d+(?:\.\d+)*)\b')

def extract_sign_code_fallback(raw: str) -> str:
    m = FALLBACK_CODE_RE.search(raw or "")
    return m.group(1).strip() if m else ""


# =========================================================
# Lock/Unlock helpers
# =========================================================
def lock_cell(ws, cell_ref: str):
    ws[cell_ref].protection = Protection(locked=True)

def unlock_cell(ws, cell_ref: str):
    ws[cell_ref].protection = Protection(locked=False)

def lock_all_cells(ws, max_row: int, max_col: int) -> None:
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).protection = Protection(locked=True)


# =========================================================
# Footer row height capture/restore
# =========================================================
def capture_row_heights(ws, start_row: int, end_row: int) -> dict:
    heights = {}
    for r in range(start_row, end_row + 1):
        heights[r] = ws.row_dimensions[r].height
    return heights

def restore_row_heights(ws, heights: dict, row_offset: int):
    for original_row, height in heights.items():
        target_row = original_row + row_offset
        ws.row_dimensions[target_row].height = height


# =========================================================
# Sign type + summary split (Excel display only)
# JSON keeps sign_type EXACTLY as PDF.
# =========================================================
TYPE_DESC_SPLIT_RE = re.compile(r"(?:\s+[-–—]\s*|\s*[-–—]\s+)", flags=re.UNICODE)

def looks_like_sign_code(code: str) -> bool:
    if not code:
        return False
    c = code.strip()
    if not c or len(c) > 60:
        return False
    if not re.match(r"^[A-Za-z0-9]", c):
        return False
    if not re.match(r"^[A-Za-z0-9./&_ ,]+$", c):
        return False
    return True

def split_sign_type_and_summary(raw: str) -> Tuple[str, str]:
    if not raw:
        return "", ""
    s = str(raw).strip()
    if not s:
        return "", ""
    parts = TYPE_DESC_SPLIT_RE.split(s, maxsplit=1)
    if len(parts) == 2:
        code = parts[0].strip()
        summary = parts[1].strip()
        if looks_like_sign_code(code):
            return code, summary
    return "", s


# =========================================================
# Merge shifting helpers
# =========================================================
CELL_RE = re.compile(r"^([A-Z]+)(\d+)$")

def shift_cell_ref(cell_ref: str, row_offset: int) -> str:
    m = CELL_RE.match(cell_ref)
    if not m:
        return cell_ref
    col, row = m.group(1), int(m.group(2))
    return f"{col}{row + row_offset}"

def parse_range(a1_range: str) -> Tuple[str, str]:
    if ":" in a1_range:
        a, b = a1_range.split(":")
        return a, b
    return a1_range, a1_range

def split_ref(ref: str):
    m = CELL_RE.match(ref)
    if not m:
        return ref, None
    return m.group(1), int(m.group(2))

def shift_range_overlap_safe(a1_range: str, footer_start_row: int, row_offset: int) -> str:
    a, b = parse_range(a1_range)
    _, a_row = split_ref(a)
    _, b_row = split_ref(b)

    if a_row is None or b_row is None:
        return a1_range

    top = min(a_row, b_row)
    bottom = max(a_row, b_row)
    should_shift = (top >= footer_start_row) or (top < footer_start_row <= bottom)

    def apply_shift(ref2):
        col, row = split_ref(ref2)
        if row is None:
            return ref2
        if should_shift:
            row += row_offset
        return f"{col}{row}"

    new_a = apply_shift(a)
    new_b = apply_shift(b)
    return f"{new_a}:{new_b}" if ":" in a1_range else new_a

def save_merged_ranges(ws) -> List[str]:
    return [str(rng) for rng in ws.merged_cells.ranges]

def unmerge_all(ws, merges: List[str]):
    for rng in merges:
        ws.unmerge_cells(rng)

def restore_merges(ws, merges: List[str], footer_start_row: int, row_offset: int):
    for rng in merges:
        new_rng = shift_range_overlap_safe(rng, footer_start_row, row_offset)
        ws.merge_cells(new_rng)

def copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    # Ensure the destination row is not hidden
    ws.row_dimensions[dst_row].hidden = False
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.font = copy(src.font)


# =========================================================
# Body adjust
# =========================================================
def adjust_body_rows_preserve_footer(
    ws,
    sign_count: int,
    body_start: int = 28,
    body_end: int = 47,
    extra_blank_rows: int = 3
) -> int:
    base_rows = body_end - body_start + 1
    needed_rows = sign_count + extra_blank_rows
    footer_start = body_end + 1
    diff = needed_rows - base_rows

    if diff == 0:
        return 0

    merges = save_merged_ranges(ws)
    unmerge_all(ws, merges)
    max_col = ws.max_column

    if diff > 0:
        ws.insert_rows(footer_start, amount=diff)
        for r in range(footer_start, footer_start + diff):
            copy_row_style(ws, src_row=body_end, dst_row=r, max_col=max_col)
    else:
        delete_count = abs(diff)
        delete_start = body_start + needed_rows
        ws.delete_rows(delete_start, amount=delete_count)

    restore_merges(ws, merges, footer_start, diff)
    return diff


# =========================================================
# Totals helpers
# =========================================================
def sum_extended(items: Optional[List[Dict[str, Any]]]) -> Optional[float]:
    if not items:
        return None
    total = 0.0
    found = False
    for it in items:
        val = safe_num(it.get("extended_total"))
        if val is not None:
            total += val
            found = True
    return total if found else None


# =========================================================
# Approx row height
# =========================================================
def approximate_autofit_rows(ws, row_start: int, row_end: int, text_cols: List[str], min_height: float = 15.0):
    CHARS_PER_LINE = 60
    LINE_HEIGHT = 15
    for r in range(row_start, row_end + 1):
        max_lines = 1
        for col in text_cols:
            v = ws[f"{col}{r}"].value
            if not v:
                continue
            text = str(v)
            explicit_lines = text.split("\n")
            line_count = 0
            for ln in explicit_lines:
                if not ln:
                    line_count += 1
                else:
                    wrapped = max(1, (len(ln) // CHARS_PER_LINE) + (1 if len(ln) % CHARS_PER_LINE else 0))
                    line_count += wrapped
            max_lines = max(max_lines, line_count)
        ws.row_dimensions[r].height = max(min_height, max_lines * LINE_HEIGHT)


# =========================================================
# Selection control
# =========================================================
def unlock_body_selection(ws, body_row_start: int, body_row_end: int) -> None:
    for r in range(body_row_start, body_row_end + 1):
        for col in ("B", "C", "D", "E"):
            unlock_cell(ws, f"{col}{r}")

    for rng in ws.merged_cells.ranges:
        min_row, max_row = rng.min_row, rng.max_row
        min_col, max_col = rng.min_col, rng.max_col
        rows_intersect = not (max_row < body_row_start or min_row > body_row_end)
        cols_intersect = not (max_col < 2 or min_col > 5)
        if rows_intersect and cols_intersect:
            ws.cell(row=min_row, column=min_col).protection = Protection(locked=False)

def apply_sheet_protection_for_selection(ws, body_row_start: int, body_row_end: int) -> None:
    max_row = max(ws.max_row, body_row_end + 5)
    max_col = max(ws.max_column, 6)
    lock_all_cells(ws, max_row=max_row, max_col=max_col)
    unlock_body_selection(ws, body_row_start, body_row_end)
    ws.protection.sheet = True


# =========================================================
# FastAPI endpoints
# =========================================================
@app.get("/health")
def health_check():
    return {
        "status": "ok",
        "template_exists": os.path.exists(TEMPLATE_PATH),
        "template_path": TEMPLATE_PATH,
        "sheet_name": SHEET_NAME,
        "logo_exists": os.path.exists(LOGO_PATH),
        "logo_path": LOGO_PATH
    }


@app.post("/generate_proposal")
def generate_proposal(payload: Dict[str, Any] = Body(default=None)):
    if not payload or "payload" not in payload:
        raise HTTPException(status_code=400, detail="Missing required field 'payload' (JSON string).")

    cleanup_old_generated_workbooks(OUTPUT_DIR, older_than_minutes=30)

    try:
        estimate_data = json.loads(payload["payload"])
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON string in 'payload': {str(e)}")

    if not isinstance(estimate_data, dict) or not estimate_data:
        raise HTTPException(status_code=400, detail="Decoded 'payload' must be a non-empty JSON object.")

    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(status_code=500, detail=f"Template not found at {TEMPLATE_PATH}")

    try:
        file_id = uuid.uuid4().hex
        out_name = f"Boyd_Proposal_{file_id}.xlsx"
        out_path = os.path.join(OUTPUT_DIR, out_name)
        generate_excel_from_data(estimate_data, out_path)
    except HTTPException:
        raise
    except Exception as e:
        logging.exception("Proposal generation failed")
        raise HTTPException(status_code=500, detail=str(e))

    base_url = os.environ.get("RAILWAY_PUBLIC_URL", "").rstrip("/")
    if not base_url:
        base_url = "https://fastapi-production-37f6.up.railway.app"

    download_url = f"{base_url}/download/{out_name}"
    return JSONResponse({"download_url": download_url, "filename": out_name})


@app.get("/download/{filename}")
def download_file(filename: str):
    file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )


# =============================================================================
# Web interface & request model
# =============================================================================
class GenerateFromPdfRequest(BaseModel):
    pdf_base64: str
    filename: str = "estimate.pdf"


@app.get("/", response_class=HTMLResponse)
async def web_interface():
    html_path = Path(__file__).parent / "templates" / "upload.html"
    if not html_path.exists():
        return """<html><body><h1>Boyd Proposal Generator</h1><p>Use POST /generate_proposal_from_pdf</p></body></html>"""
    with open(html_path) as f:
        return f.read()


# =============================================================================
# Excel generation aligned with JSON schema
# =============================================================================
def generate_excel_from_data(estimate_data: dict, output_path: str):
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(status_code=500, detail=f"Template not found at {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise HTTPException(status_code=500, detail=f"Sheet '{SHEET_NAME}' not found in workbook.")
    ws = wb[SHEET_NAME]

    ws.protection.sheet = False
    insert_logo(ws)

    FOOTER_HEIGHT_START = 48
    FOOTER_HEIGHT_END = 120
    footer_row_heights = capture_row_heights(ws, FOOTER_HEIGHT_START, FOOTER_HEIGHT_END)

    write_cell(ws, "E5", safe_str(estimate_data.get("estimate_date")))
    write_cell(ws, "D8", safe_str(estimate_data.get("project_id")))
    write_cell(ws, "C22", safe_str(estimate_data.get("salesperson")))
    write_cell(ws, "C23", safe_str(estimate_data.get("project_manager")))
    write_cell(ws, "C25", safe_str(estimate_data.get("project_description")))

    sold_to = estimate_data.get("sold_to", {}) or {}
    ship_to = estimate_data.get("ship_to", {}) or {}

    write_cell(ws, "D11", safe_str(sold_to.get("name")))
    write_cell(ws, "D13", join_address_lines(sold_to.get("address_lines") or []))
    sold_csz = " ".join([p for p in [
        safe_str(sold_to.get("city")),
        safe_str(sold_to.get("state")),
        safe_str(sold_to.get("zip"))
    ] if p.strip()])
    write_cell(ws, "D16", sold_csz)
    write_cell(ws, "D17", safe_str(sold_to.get("phone")))

    write_cell(ws, "C11", safe_str(ship_to.get("name")))
    write_cell(ws, "C13", join_address_lines(ship_to.get("address_lines") or []))
    ship_csz = " ".join([p for p in [
        safe_str(ship_to.get("city")),
        safe_str(ship_to.get("state")),
        safe_str(ship_to.get("zip"))
    ] if p.strip()])
    write_cell(ws, "C16", ship_csz)
    write_cell(ws, "C17", safe_str(ship_to.get("phone")))

    sign_types = estimate_data.get("sign_types", []) or []
    sign_count = len(sign_types)

    BODY_START = 28
    BODY_END = 47
    EXTRA_BLANK = 3

    footer_row_offset = adjust_body_rows_preserve_footer(
        ws,
        sign_count=sign_count,
        body_start=BODY_START,
        body_end=BODY_END,
        extra_blank_rows=EXTRA_BLANK
    )

    total_body_rows_needed = sign_count + EXTRA_BLANK
    body_last_row = BODY_START + total_body_rows_needed - 1

    COL_ITEM, COL_SIGN_TYPE, COL_DESC, COL_QTY, COL_UNIT, COL_TOTAL = "A", "B", "C", "D", "E", "F"
    current_row = BODY_START
    item_num = 1

    prev_priced_primary_code = None
    prev_row_was_priced_primary = False

    for sign in sign_types:
        raw_line = safe_str(sign.get("sign_type"))
        qty_val = safe_num(sign.get("qty"))
        unit_val = safe_num(sign.get("unit_price"))

        unit_cell_val = round_nearest_dollar(unit_val) if unit_val is not None else None
        is_priced_row = (qty_val is not None and unit_cell_val is not None)

        is_header_or_subtotal = is_bold_description_row(raw_line)
        is_spacer = (raw_line.strip() == "")

        ws[f"{COL_ITEM}{current_row}"].value = item_num

        # Description-only rows (headers/subtotals/spacers OR any non-priced rows)
        if (not is_priced_row) or is_header_or_subtotal or is_spacer:
            ws[f"{COL_SIGN_TYPE}{current_row}"].value = None
            ws[f"{COL_QTY}{current_row}"].value = None
            ws[f"{COL_UNIT}{current_row}"].value = None
            ws[f"{COL_TOTAL}{current_row}"].value = None
            ws[f"{COL_DESC}{current_row}"].value = raw_line

            if is_header_or_subtotal:
                ws[f"{COL_DESC}{current_row}"].font = Font(bold=True)

            # reset alternate chaining
            prev_priced_primary_code = None
            prev_row_was_priced_primary = False

            # Ensure row is visible
            ws.row_dimensions[current_row].hidden = False

            current_row += 1
            item_num += 1
            continue

        # Determine code + description
        split_code, split_desc = split_sign_type_and_summary(raw_line)

        # Prefer strict code if present; fallback to split_code or fallback extractor
        code_key = extract_sign_code_strict(raw_line) or extract_sign_code_fallback(raw_line) or (split_code.strip() if split_code else "")
        desc_summary = (split_desc.strip() if split_desc else raw_line.strip())

        # ✅ Alternate rule:
        # If two consecutive PRICED rows have the EXACT same sign type CODE (before the dash),
        # the second row becomes an Alternate (description + unit price only).
        is_alternate = (
            is_priced_row
            and bool(split_code.strip())
            and prev_row_was_priced_primary
            and prev_priced_primary_code == split_code
        )

        if is_alternate:
            # Alternate row output: ONLY description + unit price
            ws[f"{COL_SIGN_TYPE}{current_row}"].value = None
            ws[f"{COL_QTY}{current_row}"].value = None
            ws[f"{COL_TOTAL}{current_row}"].value = None
            ws[f"{COL_DESC}{current_row}"].value = f"Alternate {desc_summary}".strip()
            ws[f"{COL_UNIT}{current_row}"].value = unit_cell_val

            # alternates do not become primaries
            prev_row_was_priced_primary = False
        else:
            # Primary priced row output (normal)
            ws[f"{COL_SIGN_TYPE}{current_row}"].value = split_code if split_code else None
            ws[f"{COL_QTY}{current_row}"].value = qty_val
            ws[f"{COL_DESC}{current_row}"].value = desc_summary
            ws[f"{COL_UNIT}{current_row}"].value = unit_cell_val
            ws[f"{COL_TOTAL}{current_row}"].value = f"=D{current_row}*E{current_row}"

            # update "previous primary" tracking
            prev_priced_primary_code = split_code if split_code else None
            prev_row_was_priced_primary = bool(prev_priced_primary_code)

        # Ensure row is visible
        ws.row_dimensions[current_row].hidden = False

        current_row += 1
        item_num += 1


    # --- Totals ---
    totals = estimate_data.get("totals", {}) or {}
    grand_total = safe_num(totals.get("total"))

    shipping_total = sum_extended(estimate_data.get("shipping"))
    install_total = sum_extended(estimate_data.get("installation"))

    SUBTOTAL_CELL = "F48"
    SHIPPING_CELL = "F49"
    INSTALL_CELL = "F53"
    TOTAL_CELL = "F54"

    subtotal_cell_ref = shift_cell_ref(SUBTOTAL_CELL, footer_row_offset)
    body_sum_range = f"F{BODY_START}:F{body_last_row}"
    ws[subtotal_cell_ref].value = f"=SUM({body_sum_range})"

    write_cell(ws, shift_cell_ref(SHIPPING_CELL, footer_row_offset), shipping_total if shipping_total is not None else 0.00)
    write_cell(ws, shift_cell_ref(INSTALL_CELL, footer_row_offset), install_total if install_total is not None else 0.00)

    if grand_total is not None:
        write_cell(ws, shift_cell_ref(TOTAL_CELL, footer_row_offset), grand_total)

    # --- Layout / protection ---
    last_used_row = ws.max_row
    approximate_autofit_rows(ws, row_start=27, row_end=last_used_row, text_cols=["C"], min_height=15.0)
    restore_row_heights(ws, footer_row_heights, footer_row_offset)
    apply_sheet_protection_for_selection(ws, body_row_start=BODY_START, body_row_end=body_last_row)

    wb.save(output_path)


# =============================================================================
# PDF parsing (with subsection headers and per-subsection subtotals injected)
# =============================================================================
LINE_ITEM_RE = re.compile(
    r'^\s*(.+?)\s+(\d+(?:\.\d+)?)\s+\$?\s*([\d,]+(?:\.\d+)?)\s*\$?\s*([\d,]+(?:\.\d+)?)\s*$'
)

MONEY_RE = re.compile(r'\$?\s*([\d,]+(?:\.\d+)?)')

SUBTOTAL_RE = re.compile(r'\bSUB[- ]?TOTAL\b', re.I)
MARKUP_RE = re.compile(r'\bMARK[- ]?UP\b', re.I)
TOTAL_RE = re.compile(r'^\s*TOTAL\b', re.I)

SIGN_TABLE_HEADER_RE = re.compile(r'^\s*Sign\s*Type\s*&\s*Overall', re.I)
SIGN_PACKAGE_HEADER_RE = re.compile(r'\bPACKAGE\s*$', re.I)  # Match any line ending with PACKAGE

SHIPPING_HEADER_RE = re.compile(r'\bSHIPPING\b', re.I)
INSTALL_HEADER_RE = re.compile(r'\bINSTALL(?:ATION)?\b', re.I)
PERMIT_HEADER_RE = re.compile(r'\bPERMIT(?:TING)?\b', re.I)

DATE_RE = re.compile(r'\bDate\s+(\d{1,2}/\d{1,2}/\d{2,4})\b', re.I)
PROJECT_ID_RE = re.compile(r'\bProject\s+Id\s*:\s*(\S+)\b', re.I)
PROJECT_DESC_RE = re.compile(r'\bProject\s+Desc\.?\s*:\s*(.+?)\s*(?:Ship\s+Via|$)', re.I)

SALESPERSON_RE = re.compile(r'\bSalesperson\s*:\s*(.+)$', re.I)
SALES_REP_RE = re.compile(r'\bSales\s+Rep\s*:\s*(.+)$', re.I)

PM_RE = re.compile(r'\bProject\s+Manager\s*:\s*(.+)$', re.I)
PE_RE = re.compile(r'\bProject\s+Engineer\s*:\s*(.+)$', re.I)

CITY_STATE_ZIP_RE = re.compile(r'([A-Za-z .\'&/-]+)\s*,\s*([A-Z]{2})\s*(\d{5})(?:-\d{4})?')


def _init_data() -> Dict[str, Any]:
    return {
        "estimate_date": "",
        "project_id": "",
        "project_description": "",
        "salesperson": "",
        "project_manager": "",
        "project_engineer": "",
        "sold_to": {"name": "", "address_lines": [], "city": "", "state": "", "zip": "", "phone": ""},
        "ship_to": {"name": "", "address_lines": [], "city": "", "state": "", "zip": "", "phone": ""},
        "sign_types": [],
        "shipping": [],
        "installation": [],
        "permitting": [],
        "totals": {"sub_total": None, "mark_up": None, "total": None}
    }

def _extract_money(line: str) -> Optional[float]:
    m = MONEY_RE.search(line or "")
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except Exception:
        return None

def _clean_person_value(raw: str) -> str:
    if not raw:
        return ""
    s = raw.strip()
    for stop in [
        "Project Manager:", "Project Engineer:", "REQUESTED BY:",
        "Project Id", "Project Desc", "Ship Via", "Terms", "P.O. Number",
        "Sign Type &", "Estimate", "Date"
    ]:
        idx = s.find(stop)
        if idx != -1:
            s = s[:idx].strip()
    return s

def _apply_header_extraction(data: Dict[str, Any], line: str):
    if not data["estimate_date"]:
        m = DATE_RE.search(line)
        if m:
            data["estimate_date"] = m.group(1)

    if not data["project_id"]:
        m = PROJECT_ID_RE.search(line)
        if m:
            data["project_id"] = m.group(1)

    if not data["project_description"]:
        m = PROJECT_DESC_RE.search(line)
        if m:
            data["project_description"] = m.group(1).strip()

    if not data["salesperson"]:
        m = SALESPERSON_RE.search(line)
        if m:
            data["salesperson"] = _clean_person_value(m.group(1))
        else:
            m2 = SALES_REP_RE.search(line)
            if m2:
                data["salesperson"] = _clean_person_value(m2.group(1))

    if not data["project_manager"]:
        m = PM_RE.search(line)
        if m:
            data["project_manager"] = _clean_person_value(m.group(1))

    if not data["project_engineer"]:
        m = PE_RE.search(line)
        if m:
            data["project_engineer"] = _clean_person_value(m.group(1))

def _apply_totals_extraction(data: Dict[str, Any], line: str):
    if data["totals"]["sub_total"] is None and SUBTOTAL_RE.search(line):
        data["totals"]["sub_total"] = _extract_money(line)

    if data["totals"]["mark_up"] is None and MARKUP_RE.search(line):
        data["totals"]["mark_up"] = _extract_money(line)

    if data["totals"]["total"] is None and TOTAL_RE.search(line):
        data["totals"]["total"] = _extract_money(line)

def _normalize_merged_description(description: str) -> str:
    """
    If PDF text extraction merges two sign lines into one description, it can look like:
      "B.5.1 - ... E.1 - Maximum Occupancy (As Spec'd)"
    while qty/unit/extended belong to the E.1 line.

    We fix this by finding the LAST strict sign-code token in the description
    and slicing description from that token onward.
    """
    if not description:
        return description

    matches = list(SIGN_CODE_STRICT_ANYWHERE_RE.finditer(description))
    if not matches:
        return description

    last = matches[-1]
    start = last.start()
    # Only slice if the last token isn't already at the beginning
    if start > 0:
        return description[start:].strip()

    return description

def _parse_line_item(line: str) -> Optional[Dict[str, Any]]:
    if not line or not line.strip():
        return None
    if "$" not in line:
        return None

    m = LINE_ITEM_RE.match(line)
    if not m:
        return None

    description = m.group(1).strip()
    description = _normalize_merged_description(description)

    qty_raw = m.group(2).strip()
    unit_raw = m.group(3).strip().replace(",", "")
    ext_raw = m.group(4).strip().replace(",", "")

    qty = safe_num(qty_raw)
    if qty is not None and abs(qty - int(qty)) < 1e-9:
        qty = int(qty)

    unit_price = safe_num(unit_raw)
    extended_total = safe_num(ext_raw)

    return {"description": description, "qty": qty, "unit_price": unit_price, "extended_total": extended_total}

def _split_two_columns_by_street_numbers(line: str) -> Tuple[str, str]:
    if not line:
        return "", ""
    matches = list(re.finditer(r'\b\d{1,5}\s', line))
    if len(matches) >= 2:
        cut = matches[1].start()
        return line[:cut].strip(), line[cut:].strip()
    return line.strip(), ""

def _apply_sold_ship_extraction(data: Dict[str, Any], lines: List[str]):
    if not lines:
        return

    idx = None
    sold_name = ship_name = ""
    for i, raw in enumerate(lines[:60]):
        line = raw.strip()
        m = re.search(r'\bTo:\s*(.*?)\s+Ship\s+To:\s*(.+)$', line, re.I)
        if m:
            sold_name = m.group(1).strip()
            ship_name = m.group(2).strip()
            idx = i
            break
    if idx is None:
        return

    if sold_name and not data["sold_to"].get("name"):
        data["sold_to"]["name"] = sold_name
    if ship_name and not data["ship_to"].get("name"):
        data["ship_to"]["name"] = ship_name

    if idx + 1 < len(lines):
        addr_line = lines[idx + 1].strip()
        left_addr, right_addr = _split_two_columns_by_street_numbers(addr_line)

        if left_addr:
            al = data["sold_to"].get("address_lines", []) or []
            if left_addr not in al:
                al.append(left_addr)
            data["sold_to"]["address_lines"] = al

        if right_addr:
            al = data["ship_to"].get("address_lines", []) or []
            if right_addr not in al:
                al.append(right_addr)
            data["ship_to"]["address_lines"] = al

    if idx + 2 < len(lines):
        csz_line = lines[idx + 2].strip()
        csz = CITY_STATE_ZIP_RE.findall(csz_line)
        if csz:
            if len(csz) >= 1:
                city, state, z = csz[0]
                if not data["sold_to"].get("city"): data["sold_to"]["city"] = city.strip()
                if not data["sold_to"].get("state"): data["sold_to"]["state"] = state.strip()
                if not data["sold_to"].get("zip"): data["sold_to"]["zip"] = z.strip()
            if len(csz) >= 2:
                city, state, z = csz[1]
                if not data["ship_to"].get("city"): data["ship_to"]["city"] = city.strip()
                if not data["ship_to"].get("state"): data["ship_to"]["state"] = state.strip()
                if not data["ship_to"].get("zip"): data["ship_to"]["zip"] = z.strip()

    for raw in lines[idx:idx + 20]:
        line = raw.strip()
        if re.search(r'\bPhone\b', line, re.I):
            phone = re.sub(r'^\s*Phone\s*', '', line, flags=re.I).strip()
            if phone and not data["sold_to"].get("phone"):
                data["sold_to"]["phone"] = phone
            break

def _append_sign_row_text_only(data: Dict[str, Any], text: str):
    """
    Add a row that only populates the description column in Excel.
    IMPORTANT: allow truly blank spacer rows (text == "").
    """
    if text is None:
        return
    data["sign_types"].append({
        "sign_type": str(text),
        "qty": None,
        "unit_price": None,
        "extended_total": None
    })

def _fmt_money(v: float) -> str:
    return f"{float(v):,.2f}"

def _flush_subsection_subtotals_into_sign_types(
    data: Dict[str, Any],
    ship_sub: float,
    inst_sub: float,
    perm_sub: float,
    add_blank_row_after: bool = False
):
    def is_effectively_zero(v: float) -> bool:
        return v is None or abs(float(v)) < SUBTOTAL_EPSILON

    emitted_any = False

    if not is_effectively_zero(ship_sub):
        _append_sign_row_text_only(data, f"Shipping Subtotal ${_fmt_money(ship_sub)}")
        emitted_any = True

    if not is_effectively_zero(inst_sub):
        _append_sign_row_text_only(data, f"Install Subtotal ${_fmt_money(inst_sub)}")
        emitted_any = True

    if not is_effectively_zero(perm_sub):
        _append_sign_row_text_only(data, f"Permitting Subtotal ${_fmt_money(perm_sub)}")
        emitted_any = True

    # Spacer between subtotals and next subsection header
    if add_blank_row_after and emitted_any:
        _append_sign_row_text_only(data, "")

def _section_from_line(current: str, line: str) -> str:
    if SUBTOTAL_RE.search(line) or TOTAL_RE.search(line):
        return "totals"

    # Check for sign table header FIRST - this overrides everything
    if SIGN_TABLE_HEADER_RE.search(line):
        return "signage"

    if SHIPPING_HEADER_RE.search(line) and not SIGN_TABLE_HEADER_RE.search(line):
        return "shipping"
    if INSTALL_HEADER_RE.search(line) and not SIGN_TABLE_HEADER_RE.search(line):
        return "installation"
    if PERMIT_HEADER_RE.search(line) and not SIGN_TABLE_HEADER_RE.search(line):
        return "permitting"

    return current


def parse_boyd_estimate_from_text(text: str) -> dict:
    data = _init_data()
    section = "other"

    ship_sub = 0.0
    inst_sub = 0.0
    perm_sub = 0.0

    last_sign_header = ""

    lines = [line for line in (text or "").split("\n")]
    _apply_sold_ship_extraction(data, lines)

    for raw in lines:
        line = raw.strip("\r")

        _apply_header_extraction(data, line)
        _apply_totals_extraction(data, line)

        if SIGN_PACKAGE_HEADER_RE.search(line.strip()):
            if last_sign_header:
                _flush_subsection_subtotals_into_sign_types(
                    data, ship_sub, inst_sub, perm_sub,
                    add_blank_row_after=True
                )
            ship_sub = inst_sub = perm_sub = 0.0

            hdr = line.strip()
            if hdr and hdr != last_sign_header:
                _append_sign_row_text_only(data, hdr)
                last_sign_header = hdr

            section = "signage"
            continue

        section = _section_from_line(section, line)

        parsed = _parse_line_item(line)
        if not parsed:
            continue

        ext_val = parsed.get("extended_total") or 0.0

        if section == "signage":
            data["sign_types"].append({
                "sign_type": parsed["description"],
                "qty": parsed["qty"],
                "unit_price": parsed["unit_price"],
                "extended_total": parsed["extended_total"],
            })
        elif section == "shipping":
            data["shipping"].append({
                "description": parsed["description"],
                "qty": parsed["qty"],
                "unit_price": parsed["unit_price"],
                "extended_total": parsed["extended_total"],
                "notes": ""
            })
            ship_sub += float(ext_val)
        elif section == "installation":
            data["installation"].append({
                "description": parsed["description"],
                "qty": parsed["qty"],
                "unit_price": parsed["unit_price"],
                "extended_total": parsed["extended_total"],
                "notes": ""
            })
            inst_sub += float(ext_val)
        elif section == "permitting":
            data["permitting"].append({
                "description": parsed["description"],
                "qty": parsed["qty"],
                "unit_price": parsed["unit_price"],
                "extended_total": parsed["extended_total"],
                "notes": ""
            })
            perm_sub += float(ext_val)

    # EOF flush for final subsection
    if last_sign_header:
        _flush_subsection_subtotals_into_sign_types(
            data, ship_sub, inst_sub, perm_sub,
            add_blank_row_after=False
        )

    return data


def parse_boyd_estimate_streaming(reader: PdfReader) -> dict:
    data = _init_data()
    section = "other"

    ship_sub = 0.0
    inst_sub = 0.0
    perm_sub = 0.0

    last_sign_header = ""

    for page_index, page in enumerate(reader.pages):
        page_text = page.extract_text() or ""
        page_lines = page_text.split("\n")

        if page_index == 0:
            _apply_sold_ship_extraction(data, page_lines)

        for raw in page_lines:
            line = raw.strip("\r")

            _apply_header_extraction(data, line)
            _apply_totals_extraction(data, line)

            if SIGN_PACKAGE_HEADER_RE.search(line.strip()):
                if last_sign_header:
                    _flush_subsection_subtotals_into_sign_types(
                        data, ship_sub, inst_sub, perm_sub,
                        add_blank_row_after=True
                    )
                ship_sub = inst_sub = perm_sub = 0.0

                hdr = line.strip()
                if hdr and hdr != last_sign_header:
                    _append_sign_row_text_only(data, hdr)
                    last_sign_header = hdr

                section = "signage"
                continue

            section = _section_from_line(section, line)

            parsed = _parse_line_item(line)
            if not parsed:
                continue

            ext_val = parsed.get("extended_total") or 0.0

            if section == "signage":
                data["sign_types"].append({
                    "sign_type": parsed["description"],
                    "qty": parsed["qty"],
                    "unit_price": parsed["unit_price"],
                    "extended_total": parsed["extended_total"],
                })
            elif section == "shipping":
                data["shipping"].append({
                    "description": parsed["description"],
                    "qty": parsed["qty"],
                    "unit_price": parsed["unit_price"],
                    "extended_total": parsed["extended_total"],
                    "notes": ""
                })
                ship_sub += float(ext_val)
            elif section == "installation":
                data["installation"].append({
                    "description": parsed["description"],
                    "qty": parsed["qty"],
                    "unit_price": parsed["unit_price"],
                    "extended_total": parsed["extended_total"],
                    "notes": ""
                })
                inst_sub += float(ext_val)
            elif section == "permitting":
                data["permitting"].append({
                    "description": parsed["description"],
                    "qty": parsed["qty"],
                    "unit_price": parsed["unit_price"],
                    "extended_total": parsed["extended_total"],
                    "notes": ""
                })
                perm_sub += float(ext_val)

    # EOF flush for final subsection
    if last_sign_header:
        _flush_subsection_subtotals_into_sign_types(
            data, ship_sub, inst_sub, perm_sub,
            add_blank_row_after=False
        )

    return data


# =============================================================================
# PDF upload endpoint
# =============================================================================
@app.post("/generate_proposal_from_pdf")
async def generate_proposal_from_pdf(request: GenerateFromPdfRequest):
    try:
        logging.info("Processing PDF: %s", request.filename)

        pdf_bytes = base64.b64decode(request.pdf_base64)
        pdf_file = BytesIO(pdf_bytes)
        reader = PdfReader(pdf_file)

        num_pages = len(reader.pages)
        logging.info("PDF pages: %d", num_pages)

        use_streaming = num_pages >= MAX_PDF_PAGES_STREAMING
        estimate_data = None

        if use_streaming:
            logging.info("Using streaming parser (page-by-page) due to page count.")
            estimate_data = parse_boyd_estimate_streaming(reader)
        else:
            text_parts = []
            total_chars = 0
            for page in reader.pages:
                t = page.extract_text() or ""
                total_chars += len(t)
                text_parts.append(t)
                if total_chars > MAX_EXTRACTED_TEXT_CHARS:
                    logging.info(
                        "Extracted text exceeds %d chars; switching to streaming fallback.",
                        MAX_EXTRACTED_TEXT_CHARS
                    )
                    estimate_data = parse_boyd_estimate_streaming(reader)
                    break

            if estimate_data is None:
                text = "\n".join(text_parts) + "\n"
                estimate_data = parse_boyd_estimate_from_text(text)

        items_count = len(estimate_data.get("sign_types", []))
        logging.info("Parsed sign line items (including subsection headers and subtotals): %d", items_count)

        if items_count == 0:
            raise HTTPException(
                status_code=400,
                detail="No sign line items found in PDF (could not detect sign section or parse price rows)."
            )

        cleanup_old_generated_workbooks(OUTPUT_DIR)

        unique_id = uuid.uuid4().hex
        output_filename = f"Boyd_Proposal_{unique_id}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        generate_excel_from_data(estimate_data, output_path)

        base_url = os.environ.get("RAILWAY_PUBLIC_URL", "").rstrip("/")
        if not base_url:
            base_url = "https://fastapi-production-37f6.up.railway.app"

        download_url = f"{base_url}/download/{output_filename}"

        return {
            "status": "success",
            "filename": output_filename,
            "download_url": download_url,
            "items_count": items_count,
            "project_id": estimate_data.get("project_id", ""),
            "total": estimate_data.get("totals", {}).get("total", None),
            "used_streaming_fallback": bool(use_streaming)
        }

    except HTTPException:
        raise
    except Exception as e:
        logging.error("Error processing PDF: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error processing PDF: {str(e)}")